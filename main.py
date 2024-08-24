import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
import os

# Define the path to the Excel file
file_path = '/Users/zeeshanwaheed/Desktop/Melon Pan/Orders.xlsx'


# Parse the order string
def parse_order_string(order_str):
    order_data = {}
    lines = order_str.strip().split('\n')
    for line in lines:
        if ':' in line:
            key, value = line.split(':', 1)
            order_data[key.strip()] = value.strip()
    return order_data


# Create or update the Excel file
def update_excel(file_path, order_data):
    if os.path.exists(file_path):
        # Load existing workbook
        book = load_workbook(file_path)
        sheet = book.active
    else:
        # Create new workbook
        book = Workbook()
        sheet = book.active
        # Create the header row
        sheet.append(['Name', 'Address', 'Contact', 'Order Detail', 'Delivery Status'])

    # Append the new data
    new_data = {
        'Name': order_data.get('Name', ''),
        'Address': order_data.get('Address', ''),
        'Contact': order_data.get('Contact', ''),
        'Order Detail': order_data.get('Order Detail', ''),
        'Delivery Status': 'Order Received'
    }

    # Add the new data row
    row_num = sheet.max_row + 1
    sheet.append([new_data['Name'], new_data['Address'], new_data['Contact'], new_data['Order Detail'],
                  new_data['Delivery Status']])

    # Apply font size to cells in the new row
    for col in range(1, 6):  # Columns A to E
        cell = sheet.cell(row=row_num, column=col)
        cell.font = Font(size=14)

    # Define dropdown menu
    dv = DataValidation(type="list", formula1='"Order Received,Processing,Delivered"', showDropDown=True)
    sheet.add_data_validation(dv)

    # Apply the data validation to the "Delivery Status" column
    dv.add(sheet[f"E{row_num}"])

    # Define background colors
    colors = {
        'Order Received': 'FFFF00',  # Yellow
        'processing': '0000FF',  # Blue
        'delivered': '00FF00'  # Green
    }

    # Apply the colors to cells in the Delivery Status column
    for cell in sheet['E']:
        if cell.row == 1:
            continue  # Skip header
        status = cell.value
        if status in colors:
            color_code = colors[status]
            cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

    # Save the file
    book.save(file_path)


# Main loop
def main():
    while True:
        user_input = input("Do you want to add an order? Type '0' to exit or press Enter to add an order: ")

        if user_input == '0':
            print("Exiting...")
            break

        print("Please paste the order details (press Enter on an empty line to finish):")

        # Capture multiline input
        order_lines = []
        while True:
            line = input()
            if not line:  # Break on empty line
                break
            order_lines.append(line)

        order_text = '\n'.join(order_lines)

        # Parse the order string
        order_data = parse_order_string(order_text)

        # Check if name is extracted properly
        name = order_data.get('Name', 'Unknown')
        if name == 'Unknown':
            print("Could not extract order details. Please make sure the format is correct.")
            continue

        # Print confirmation message
        print(f"Order for {name} added.")

        # Update the Excel file
        update_excel(file_path, order_data)

        # Prompt to add another order or exit
        print()  # Print a blank line for better readability


if __name__ == "__main__":
    main()
